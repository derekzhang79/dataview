from pymongo import MongoClient, ASCENDING, DESCENDING
from pymongo import MongoClient, ASCENDING, DESCENDING
import pandas as pd
import argparse
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

def highlight_min_values_in_excel(df, output_file):
    """比较Excel文件中bidprice9、price和bidprice10三列并将最小值标红；如有0值则忽略并比较另外两列；确保每行最多只有一个标红"""
    # 使用ExcelWriter和openpyxl引擎
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 先将数据写入Excel
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        
        # 获取工作表
        worksheet = writer.sheets['Sheet1']
        
        # 获取列名行索引（第1行）
        header_row = 1
        
        # 初始化目标列索引
        bidprice9_col = None
        price_col = None
        bidprice10_col = None
        
        # 查找目标列的索引
        for cell in worksheet[header_row]:
            if cell.value == 'bidprice9':
                bidprice9_col = cell.column
            elif cell.value == 'price':
                price_col = cell.column
            elif cell.value == 'bidprice10':
                bidprice10_col = cell.column
        
        # 创建红色字体样式
        red_font = Font(color='FF0000')
        
        # 检查是否找到了所有需要的列
        if bidprice9_col and price_col and bidprice10_col:
            # 从第2行开始遍历数据行（跳过表头）
            for row in range(2, worksheet.max_row + 1):
                try:
                    # 获取三列的值和单元格
                    bidprice9_cell = worksheet.cell(row=row, column=bidprice9_col)
                    price_cell = worksheet.cell(row=row, column=price_col)
                    bidprice10_cell = worksheet.cell(row=row, column=bidprice10_col)
                    
                    # 尝试将值转换为数字
                    try:
                        bidprice9_val = float(bidprice9_cell.value) if bidprice9_cell.value else None
                    except (ValueError, TypeError):
                        bidprice9_val = None
                    
                    try:
                        price_val = float(price_cell.value) if price_cell.value else None
                    except (ValueError, TypeError):
                        price_val = None
                    
                    try:
                        bidprice10_val = float(bidprice10_cell.value) if bidprice10_cell.value else None
                    except (ValueError, TypeError):
                        bidprice10_val = None
                    
                    # 创建值和单元格的映射列表，同时排除0值
                    value_cell_pairs = []
                    if bidprice9_val is not None and bidprice9_val != 0:
                        value_cell_pairs.append((bidprice9_val, bidprice9_cell))
                    if price_val is not None and price_val != 0:
                        value_cell_pairs.append((price_val, price_cell))
                    if bidprice10_val is not None and bidprice10_val != 0:
                        value_cell_pairs.append((bidprice10_val, bidprice10_cell))
                    
                    # 如果有至少两个有效的非0值进行比较
                    if len(value_cell_pairs) >= 2:
                        # 找出最小值及其对应的单元格
                        min_val, min_cell = min(value_cell_pairs, key=lambda x: x[0])
                        # 将最小值标红
                        min_cell.font = red_font
                except Exception:
                    # 安全处理任何异常，确保程序继续运行
                    continue
        else:
            # 如果缺少必要的列，打印警告信息
            missing_cols = []
            if not bidprice9_col: missing_cols.append('bidprice9')
            if not price_col: missing_cols.append('price')
            if not bidprice10_col: missing_cols.append('bidprice10')
            print(f"警告: Excel文件中缺少以下必要列: {', '.join(missing_cols)}")


def export_mongodb_to_excel():
    # 创建命令行参数解析器
    parser = argparse.ArgumentParser(description='从MongoDB导出数据到Excel文件')
    parser.add_argument('--db', default='foooodata', help='MongoDB数据库名称(默认: foooodata)')
    parser.add_argument('collection', help='MongoDB集合名称(必须指定)')
    parser.add_argument('--fields', help='要导出的字段名，多个字段用逗号分隔(默认导出所有字段)')
    parser.add_argument('--sort', default='nameid', help='排序字段(默认: nameid)')
    parser.add_argument('--order', choices=['asc', 'desc'], default='asc', help='排序顺序(默认: asc升序)')
    
    # 解析命令行参数
    args = parser.parse_args()
    
    # 配置参数
    db_name = args.db
    collection_name = args.collection
    output_file = f'{collection_name}.xlsx'
    
    # 处理字段参数
    fields = None
    if args.fields:
        fields = {field.strip(): 1 for field in args.fields.split(',')}
        # 确保_id字段不被包含（除非明确指定）
        if '_id' not in fields:
            fields['_id'] = 0
    else:
        # 默认不导出level1, level2, level3, spec和_id字段
        fields = {'level1': 0, 'level2': 0, 'level3': 0, 'spec': 0, '_id': 0}
    
    # 处理排序参数
    sort_field = args.sort
    sort_order = ASCENDING if args.order == 'asc' else DESCENDING
    sort_criteria = [(sort_field, sort_order)]
    
    try:
        # 连接到MongoDB
        print(f"正在连接MongoDB数据库: {db_name}")
        print(f"正在访问集合: {collection_name}")
        client = MongoClient('mongodb://localhost:27017/')
        db = client[db_name]
        collection = db[collection_name]
        
        # 查询所有文档，应用字段筛选和排序
        print(f"正在查询文档...")
        print(f"  导出字段: {'所有字段(除level1,level2,level3,spec,_id)' if args.fields is None else args.fields}")
        print(f"  排序方式: {sort_field} {args.order}")
        cursor = collection.find({}, fields).sort(sort_criteria)
        
        # 将游标转换为列表，然后创建DataFrame
        print("正在处理数据...")
        df = pd.DataFrame(list(cursor))
        
        # 如果用户指定了字段，按照用户指定的顺序排列列
        if args.fields:
            column_order = [field.strip() for field in args.fields.split(',')]
            # 过滤出存在的列
            existing_columns = [col for col in column_order if col in df.columns]
            # 补充其他可能存在的列
            remaining_columns = [col for col in df.columns if col not in existing_columns]
            df = df[existing_columns + remaining_columns]
        
        # 检查是否有数据
        if df.empty:
            print("警告: 没有找到数据")
            return
        
        # 从constprice集合获取price数据并添加到DataFrame
        print("正在从constprice集合获取产品价格数据...")
        # 获取constprice集合引用
        constprice_collection = db['constprice']
        
        # 创建一个新的price列，初始化为0
        df['price'] = 0
        
        # 用于统计匹配情况
        matched_count = 0
        unmatched_count = 0
        
        # 遍历DataFrame，使用nameid查询constprice中的price
        for index, row in df.iterrows():
            try:
                nameid = row['nameid']
                
                # 尝试将nameid转换为整数，因为constprice集合中的nameid是整数类型
                try:
                    # 先去除可能的前导零和空格
                    if isinstance(nameid, str):
                        nameid_str = nameid.strip()
                        # 尝试转换为整数
                        query_nameid = int(nameid_str)
                    else:
                        query_nameid = int(nameid)
                except (ValueError, TypeError):
                    # 如果转换失败，使用原始nameid继续查询
                    query_nameid = nameid
                    
                # 查询constprice集合中对应nameid的文档
                const_price_doc = constprice_collection.find_one({'nameid': query_nameid}, {'price': 1, '_id': 0})
                
                # 如果找到对应文档，处理price值
                if const_price_doc and 'price' in const_price_doc:
                    price_value = const_price_doc['price']
                    matched_count += 1
                    
                    # 处理不同类型的price值
                    try:
                        # 尝试将price值转换为浮点数
                        if isinstance(price_value, str):
                            # 处理字符串类型的price
                            price_value = price_value.strip().upper()
                            if price_value in ('N/A', 'NA', '', 'NONE', 'NULL'):
                                df.at[index, 'price'] = 0
                            else:
                                # 尝试将字符串转换为数字
                                df.at[index, 'price'] = float(price_value)
                        elif isinstance(price_value, (int, float)):
                            # 对于数字类型的price，直接使用
                            df.at[index, 'price'] = price_value
                        else:
                            # 其他类型默认为0
                            df.at[index, 'price'] = 0
                    except (ValueError, TypeError):
                        # 如果转换失败，设置为0
                        df.at[index, 'price'] = 0
                else:
                    unmatched_count += 1
                    # 如果没有找到对应文档，保持price为0
                    df.at[index, 'price'] = 0
                    # 可选：添加日志记录未匹配的nameid
                    # print(f"未找到nameid为{nameid}的产品价格数据")
            except Exception as e:
                # 安全地处理nameid可能不存在的情况
                try:
                    error_nameid = row.get('nameid', '未知')
                    print(f"获取产品ID为{error_nameid}的价格时出错: {str(e)}")
                except:
                    print(f"获取产品价格时出错: {str(e)}")
                # 即使出错，也要确保price列为0
                df.at[index, 'price'] = 0
                continue
        
        # 显示匹配统计信息
        print(f"从constprice集合获取价格数据完成：")
        print(f"  成功匹配: {matched_count} 条记录")
        print(f"  未找到匹配: {unmatched_count} 条记录")
        
        # 重新排列列的顺序，确保符合要求
        required_columns = ['name', 'nameid', 'number9', 'price9', 'bidprice9', 'price', 'bidprice10', 'price10', 'number10']
        # 确保DataFrame中存在的列按照指定顺序排列
        existing_required_columns = [col for col in required_columns if col in df.columns]
        # 补充其他可能存在的列
        remaining_columns = [col for col in df.columns if col not in existing_required_columns]
        # 重新排列DataFrame的列
        df = df[existing_required_columns + remaining_columns]
        
        # 导出DataFrame到Excel文件并高亮显示较小值
        print(f"正在导出到Excel文件: {output_file}")
        print(f"  导出列: {', '.join(df.columns.tolist())}")
        print(f"  正在处理数值比较，将较小值标记为红色字体")
        highlight_min_values_in_excel(df, output_file)
        
        print(f"导出完成! 共导出{len(df)}条记录")
        print(f"输出文件: {output_file}")
        
        client.close()
        
    except pymongo.errors.ServerSelectionTimeoutError:
        print("错误: 无法连接到MongoDB，请确保MongoDB服务正在运行")
    except Exception as e:
        print(f"错误: {str(e)}")

if __name__ == "__main__":
    export_mongodb_to_excel()